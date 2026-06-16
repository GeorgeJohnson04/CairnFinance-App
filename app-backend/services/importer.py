"""Import 'Finance Project.xlsx' into a user's portfolio — with the
workbook's known data errors corrected on the way in:

1. Account names are trimmed ('Fidelity ' → 'Fidelity'), so aggregations
   can never silently drop rows the way SUMPRODUCT did.
2. Asset types are normalized against the Holdings sheet (the trade log
   had XLE marked 'Stock' though it's an ETF).
3. Money-market trades with a missing price default to $1.00/share.
4. Realized P&L is recomputed consistently for every sell.
5. Every account (including Robinhood Roth IRA, which the Excel dashboard
   omitted) is part of every breakdown, because breakdowns are grouped
   from the data.
6. Roth IRA contributions and the annual limit come in as structured
   records instead of hardcoded cells.
"""
import re
from datetime import date, datetime

from openpyxl import load_workbook

ASSET_TYPES = {"Stock", "ETF", "Crypto", "Options", "Cash"}
INCOME_TYPES = {"Interest", "Qualified Dividend", "Ordinary Dividend", "Other"}


def _account_kind(name: str) -> str:
    low = name.lower()
    if "roth" in low or "ira" in low or "401" in low:
        return "retirement"
    if "cma" in low or "cash" in low or "checking" in low:
        return "cash"
    if "saving" in low:
        return "savings"
    return "brokerage"


def _as_date(value) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip()[:10],
                                     "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            return None
    return None


def _num(value) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None


def import_workbook(db, user_id: int, path: str, replace: bool = False):
    """Returns (counts dict, list of correction notes)."""
    wb = load_workbook(path, data_only=True)
    notes: list[str] = []
    counts = {"accounts": 0, "holdings": 0, "trades": 0,
              "incomes": 0, "contributions": 0}

    if replace:
        for table in ("contributions", "incomes", "trades", "holdings",
                      "accounts", "goals"):
            db.execute(f"DELETE FROM {table} WHERE user_id = ?", (user_id,))

    # ------------------------------------------------------------ accounts
    account_ids: dict[str, int] = {}

    def ensure_account(raw_name) -> int | None:
        if not raw_name or not str(raw_name).strip():
            return None
        name = re.sub(r"\s+", " ", str(raw_name)).strip()
        if name != str(raw_name):
            note = f"Normalized account name {str(raw_name)!r} → {name!r}"
            if note not in notes:
                notes.append(note)
        if name in account_ids:
            return account_ids[name]
        row = db.execute(
            "SELECT id FROM accounts WHERE user_id = ? AND name = ?",
            (user_id, name)).fetchone()
        if row:
            account_ids[name] = row["id"]
        else:
            cur = db.execute(
                "INSERT INTO accounts (user_id, name, kind) VALUES (?, ?, ?)",
                (user_id, name, _account_kind(name)))
            account_ids[name] = cur.lastrowid
            counts["accounts"] += 1
        return account_ids[name]

    # ------------------------------------------------------------ holdings
    ticker_types: dict[str, str] = {}   # authoritative asset type per ticker
    ticker_names: dict[str, str] = {}
    if "Holdings" in wb.sheetnames:
        ws = wb["Holdings"]
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row,
                                max_col=12, values_only=True):
            account, asset_type, ticker, name = row[0], row[1], row[2], row[3]
            qty, avg_cost, price, industry = row[4], row[5], row[6], row[11]
            if not ticker or not str(ticker).strip():
                continue
            account_id = ensure_account(account)
            if account_id is None:
                continue
            ticker = str(ticker).strip().upper()
            asset_type = str(asset_type or "Stock").strip().title()
            asset_type = "ETF" if asset_type == "Etf" else asset_type
            if asset_type not in ASSET_TYPES:
                asset_type = "Stock"
            ticker_types[ticker] = asset_type
            ticker_names[ticker] = str(name or ticker).strip()
            db.execute(
                "INSERT INTO holdings (user_id, account_id, asset_type,"
                " ticker, name, industry, quantity, avg_cost, current_price,"
                " price_updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?,"
                " datetime('now'))"
                " ON CONFLICT (user_id, account_id, ticker, asset_type)"
                " DO UPDATE SET quantity = excluded.quantity,"
                " avg_cost = excluded.avg_cost,"
                " current_price = excluded.current_price,"
                " name = excluded.name, industry = excluded.industry",
                (user_id, account_id, asset_type, ticker,
                 str(name or ticker).strip(), str(industry or "").strip(),
                 _num(qty) or 0.0, _num(avg_cost) or 0.0, _num(price) or 0.0),
            )
            counts["holdings"] += 1

    # -------------------------------------------------------------- trades
    if "Trade Log" in wb.sheetnames:
        ws = wb["Trade Log"]
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row,
                                max_col=12, values_only=True):
            (raw_date, account, asset_type, ticker, action, qty, price,
             _total, cost_basis, *_rest) = row
            trade_date = _as_date(raw_date)
            if not trade_date or not ticker or not action:
                continue
            account_id = ensure_account(account)
            if account_id is None:
                continue
            ticker = str(ticker).strip().upper()
            action = str(action).strip().title()
            if action not in ("Buy", "Sell"):
                continue
            asset_type = str(asset_type or "Stock").strip().title()
            asset_type = "ETF" if asset_type == "Etf" else asset_type
            if ticker in ticker_types and ticker_types[ticker] != asset_type:
                notes.append(
                    f"Corrected asset type for {ticker} trade on "
                    f"{trade_date}: {asset_type} → {ticker_types[ticker]}")
                asset_type = ticker_types[ticker]
            if asset_type not in ASSET_TYPES:
                asset_type = "Stock"

            quantity = _num(qty)
            unit_price = _num(price)
            if quantity is None or quantity <= 0:
                continue
            if unit_price is None:
                if asset_type == "Cash":
                    unit_price = 1.0
                    notes.append(
                        f"Filled missing price for {ticker} "
                        f"{action.lower()} on {trade_date} ($1.00/share "
                        "money market)")
                else:
                    continue

            basis = _num(cost_basis)
            realized = None
            if action == "Sell":
                if asset_type == "Cash":
                    basis = 1.0
                    realized = 0.0
                elif basis is not None:
                    realized = (unit_price - basis) * quantity

            db.execute(
                "INSERT INTO trades (user_id, account_id, trade_date,"
                " asset_type, ticker, action, quantity, price, cost_basis,"
                " realized_pl, note) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?,"
                " 'Imported from Excel')",
                (user_id, account_id, trade_date, asset_type, ticker, action,
                 quantity, unit_price, basis, realized),
            )
            counts["trades"] += 1

    # -------------------------------------------------------------- income
    if "Income" in wb.sheetnames:
        ws = wb["Income"]
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row,
                                max_col=6, values_only=True):
            raw_date, account, ticker, income_type, description, amount = row
            income_date = _as_date(raw_date)
            value = _num(amount)
            if not income_date or value is None:
                continue
            account_id = ensure_account(account)
            if account_id is None:
                continue
            income_type = str(income_type or "Other").strip()
            if income_type not in INCOME_TYPES:
                income_type = "Other"
            ticker = str(ticker or "").strip().upper()
            ticker = "" if ticker in ("-", "—") else ticker
            db.execute(
                "INSERT INTO incomes (user_id, account_id, income_date,"
                " ticker, type, description, amount)"
                " VALUES (?, ?, ?, ?, ?, ?, ?)",
                (user_id, account_id, income_date, ticker, income_type,
                 str(description or "").strip(), value),
            )
            counts["incomes"] += 1

    # ----------------------------------------- Roth contributions & limit
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        contrib_year = None
        in_section = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2,
                                values_only=True):
            label = str(row[0] or "").strip()
            value = row[1]
            year_match = re.match(r"^(20\d\d).*Contribution", label)
            if year_match:
                contrib_year = int(year_match.group(1))
                in_section = True
                continue
            if not in_section:
                continue
            if re.search(r"\bLimit\b", label, re.I) and _num(value):
                db.execute("UPDATE users SET ira_limit = ? WHERE id = ?",
                           (_num(value), user_id))
                notes.append(f"Set annual IRA limit to ${_num(value):,.0f}")
            elif (label and _num(value) and "total" not in label.lower()
                  and "remaining" not in label.lower()
                  and "match" not in label.lower()
                  and "account" != label.lower()):
                account_id = ensure_account(label)
                if account_id and contrib_year:
                    db.execute(
                        "INSERT INTO contributions (user_id, account_id,"
                        " year, amount, contrib_date, note)"
                        " VALUES (?, ?, ?, ?, ?, 'Imported from Excel')",
                        (user_id, account_id, contrib_year, _num(value),
                         f"{contrib_year}-01-01"),
                    )
                    counts["contributions"] += 1

    db.commit()
    notes.append("All breakdowns now include every account — the Excel "
                 "dashboard was missing 'Robinhood Roth IRA' in its "
                 "By-Account and Income summaries.")
    notes.append("'Without retirement' totals are computed from account "
                 "type, fixing the W/O ROTH formula that excluded "
                 "Fidelity CMA.")
    return counts, notes
